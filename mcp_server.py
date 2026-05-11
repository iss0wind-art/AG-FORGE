"""
피지수(Physis) MCP 서버 — mcp_server.py
Claude Code에서 피지수 독립 뇌를 직접 호출한다.
stdio 전송 방식 (Claude Code 표준).

Claude Code settings.json 등록:
{
  "mcpServers": {
    "physis": {
      "command": "python",
      "args": ["d:/Git/AG-Forge/mcp_server.py"]
    }
  }
}

호출 방식:
  Claude Code 채팅에서 자연스럽게 "피지수야 ..." 라고 호출하면 자동 응답
  또는 @physis 명령어로 명시적 호출도 가능
"""
from __future__ import annotations
import io
import os
import re
import sys
from pathlib import Path

from dotenv import load_dotenv
load_dotenv()

# Windows cp949 환경에서 이모지 포함 출력 시 인코딩 오류 방지
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "buffer"):
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from mcp.server.fastmcp import FastMCP

# AG-Forge 루트를 sys.path에 추가 (어떤 cwd에서 실행해도 동작)
_FORGE_ROOT = Path(__file__).parent
sys.path.insert(0, str(_FORGE_ROOT))

# EXCEL_DIAT 루트 추가 (합체 엔진)
_DIAT_ROOT = Path("D:/Git/EXCEL_DIAT")
if _DIAT_ROOT.exists():
    sys.path.insert(0, str(_DIAT_ROOT))

from scripts.brain_loader import BrainResponse, LLMProvider, run
from scripts.router_agent import route
from scripts.observability import summarize_session, LOG_PATH

mcp = FastMCP("physis")

BRAIN_ROOT = _FORGE_ROOT


# ── Provider 빌드 ─────────────────────────────────────────────────────────────

class _FallbackProvider(LLMProvider):
    """Gemini API 키 없을 때 사용하는 MCP 전용 fallback."""

    def generate(self, system_instruction, context_layers, task, model, thinking_budget):
        routing = route(task)
        return BrainResponse(
            text=(
                f"[{model}] {task} → 라우팅 완료.\n"
                "실제 응답은 GEMINI_API_KEY 설정 후 활성화됩니다."
            ),
            model=model,
            task_type=routing.task_type.value,
            tokens_used=0,
            cache_hit=False,
        )


def _build_provider() -> LLMProvider:
    """Claude→Qwen→DeepSeek→Groq→Gemini 순서의 ChainedProvider 반환."""
    from scripts.brain_loader import (
        GroqProvider, DeepSeekProvider, GeminiProvider,
        QwenProvider, ClaudeProvider, ChainedProvider
    )

    providers = []
    # 2026-04-30: DeepSeek R1 최우선 (Claude 크레딧 소진)
    priority_keys = [
        ("DEEPSEEK_API_KEY", DeepSeekProvider),
        ("QWEN_API_KEY",     QwenProvider),
        ("GROQ_API_KEY",     GroqProvider),
        ("GEMINI_API_KEY",   GeminiProvider),
        ("CLAUDE_API_KEY",   ClaudeProvider),
    ]

    for key_name, cls in priority_keys:
        key = os.environ.get(key_name, "")
        if key and key.strip():
            try:
                providers.append(cls(key))
            except Exception as e:
                print(f"[physis] {key_name} 초기화 실패: {e}", file=sys.stderr)

    if providers:
        return ChainedProvider(providers)
    return _FallbackProvider()


# ── 툴 1: ask_brain ───────────────────────────────────────────────────────────

@mcp.tool()
def physis(task: str) -> str:
    """
    피지수 독립 뇌에 작업을 전달하고 응답을 반환한다.
    헌법 게이트(CONSTITUTION.md)를 통과한 응답만 반환된다.

    Claude Code에서 "피지수야" 라고 호출하면 자동으로 이 도구가 실행됨.
    예: "피지수야 이 함수 최적화해줄 수 있어?"

    Args:
        task: 수행할 작업 또는 질문 (예: "이 코드 리뷰해줘", "아키텍처 설계해줘")

    Returns:
        헌법 게이트를 통과한 피지수의 응답 텍스트
    """
    if not task.strip():
        return "[오류] task는 비어있을 수 없습니다."

    provider = _build_provider()

    try:
        result: BrainResponse = run(task.strip(), provider)
        return result.text
    except Exception as exc:
        print(f"[physis] run 실패: {type(exc).__name__}: {exc}", file=sys.stderr)
        return "[AG-Forge 오류] 뇌 응답 중 문제가 발생했습니다. 로그를 확인하세요."


# ── 툴 2: get_brain_status ────────────────────────────────────────────────────

@mcp.tool()
def physis_status() -> dict:
    """
    피지수의 현재 상태를 반환한다.
    brain.md 요약과 마지막 라우팅 정보를 포함한다.

    반환:
        피지수 상태 정보 (뇌 활성화 상황, 마지막 작업 로그)
    """
    brain_path = BRAIN_ROOT / "brain.md"
    judgment_path = BRAIN_ROOT / "judgment.md"

    try:
        content = brain_path.read_text(encoding="utf-8")
        summary_match = re.search(r"## 8\..*?```yaml(.*?)```", content, re.DOTALL)
        summary = summary_match.group(1).strip() if summary_match else content[:300]
    except FileNotFoundError:
        summary = "[brain.md 없음]"

    try:
        judgment = judgment_path.read_text(encoding="utf-8")
        log_lines = [l for l in judgment.splitlines() if "|" in l and "gemini" in l.lower()]
        last_routing = log_lines[-1].strip() if log_lines else "없음"
    except FileNotFoundError:
        last_routing = "[judgment.md 없음]"

    return {
        "brain_summary": summary,
        "active_layer": "brain.md",
        "last_routing": last_routing,
    }


# ── 툴 3: get_brain_logs ──────────────────────────────────────────────────────

@mcp.tool()
def physis_logs() -> dict:
    """
    피지수의 observability 세션 통계를 반환한다.
    총 요청 수, 누적 비용, 캐시 히트율 등.

    Returns:
        피지수 활동 리포트 (비용/캐시/성능)
    """
    return summarize_session(LOG_PATH)


# ── 툴 4: Excel Surgeon (EXCEL_DIAT 통합) ───────────────────────────────────

@mcp.tool()
def excel_surgical_diet(input_path: str, output_path: str = None) -> str:
    """
    대용량 엑셀 파일을 다이어트시킨다 (90% 이상 압축).
    노란색 셀(편집 대상)만 남기고 나머지 서식/이미지 등을 최적화한다.

    Args:
        input_path: 원본 엑셀 파일 경로 (예: "D:/Git/EXCEL_DIAT/large.xlsx")
        output_path: 결과 저장 경로 (기본값: input_path_diet.xlsx)

    Returns:
        수행 결과 메시지 및 생성된 파일 경로
    """
    try:
        from diet_engine import surgical_diet

        in_p = Path(input_path)
        if not in_p.exists():
            return f"[오류] 파일을 찾을 수 없습니다: {input_path}"

        if not output_path:
            output_path = str(in_p.parent / f"{in_p.stem}_diet{in_p.suffix}")

        success, msg = surgical_diet(input_path, output_path)
        if success:
            return f"[성공] 엑셀 수술 완료. 결과: {output_path} ({msg})"
        else:
            return f"[실패] {msg}"
    except Exception as e:
        return f"[엔진 오류] {str(e)}"


@mcp.tool()
def extract_boq_data(project_id: str) -> dict:
    """
    프로젝트 ID를 기준으로 다이어트된 엑셀에서 BOQ 내역 데이터를 추출한다.

    Args:
        project_id: EXCEL_DIAT 프로젝트 ID

    Returns:
        추출된 데이터 요약 (시트 목록, 항목 수 등)
    """
    # EXCEL_DIAT의 API 로직을 참고하여 구현 (현재는 메타데이터 반환 위주)
    try:
        project_dir = Path(f"D:/Git/EXCEL_DIAT/api/projects/{project_id}")
        diet_path = project_dir / "diet.xlsx"
        if not diet_path.exists():
            return {"status": "error", "message": f"프로젝트를 찾을 수 없습니다: {project_id}"}

        from openpyxl import load_workbook
        wb = load_workbook(str(diet_path), read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()

        return {
            "status": "success",
            "project_id": project_id,
            "sheets": sheets,
            "file_size": diet_path.stat().st_size
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}


@mcp.tool()
def generate_gabji_report(project_id: str) -> dict:
    """
    공종별 집계표(갑지) 데이터를 생성한다.

    Args:
        project_id: EXCEL_DIAT 프로젝트 ID

    Returns:
        집계표 데이터 (전회, 금회, 누적 기성 등)
    """
    try:
        from gabji_engine import build_jibgye_table
        db_path = "D:/Git/EXCEL_DIAT/boq.db"

        rows = build_jibgye_table(db_path, project_id)
        if not rows:
            return {"status": "error", "message": "집계할 데이터가 없습니다."}

        return {
            "status": "success",
            "project_id": project_id,
            "jibgye": rows
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}


# ── 단군 브리지 내부 유틸 ────────────────────────────────────────────────────

_VALID_URGENCIES = {"normal", "high", "emergency"}
_URGENCY_TIMEOUT = {"high": 120, "emergency": 30}
_DANGUN_API_URL = "http://localhost:8020/api/dangun_brain"


def _call_dangun_brain(issue: str) -> str:
    """단군 HTTP API를 호출한다 (포트 8020)."""
    import urllib.request
    import json as _json
    payload = _json.dumps({"issue": issue}).encode()
    req = urllib.request.Request(
        _DANGUN_API_URL,
        data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = _json.loads(resp.read())
            if "error" in data:
                return f"[단군/오류] {data['error']}"
            return data.get("result", "[단군/응답없음]")
    except Exception as e:
        err_msg = str(e).encode("utf-8", errors="replace").decode("utf-8")
        print(f"[physis→dangun] 단군 호출 실패: {type(e).__name__}: {err_msg}", file=sys.stderr)
        return f"[단군/연결오류] {err_msg}"


# ── 툴 7: 본영 단군 escalation (다리 B — 실제 연결) ─────────────────────────

def physis_escalate_dangun(
    issue: str,
    urgency: str,
    context: dict | None = None,
) -> dict:
    """
    피지수 → 본영 단군 escalation 요청. 단군 Python 브레인 직접 호출.

    Args:
        issue: 에스컬레이션 사안 (비어있으면 오류)
        urgency: "normal" | "high" | "emergency"
        context: 추가 맥락 딕셔너리 (기본값 {})

    Returns:
        단군 응답 또는 오류 dict
    """
    if context is None:
        context = {}

    if not issue.strip():
        return {"status": "error", "message": "issue는 비어있을 수 없습니다."}

    if urgency not in _VALID_URGENCIES:
        return {
            "status": "error",
            "message": f"urgency는 {_VALID_URGENCIES} 중 하나여야 합니다.",
        }

    if urgency == "normal":
        return {
            "status": "error",
            "message": "paperclip 미가동 상태에서는 normal urgency 처리를 할 수 없습니다.",
        }

    prompt = f"[{urgency.upper()} 에스컬레이션] {issue.strip()}"
    if context:
        import json as _json
        prompt += f"\n맥락: {_json.dumps(context, ensure_ascii=False)}"

    response = _call_dangun_brain(prompt)
    return {
        "status": "escalated",
        "urgency": urgency,
        "timeout_sec": _URGENCY_TIMEOUT[urgency],
        "context": context,
        "dangun_response": response,
    }


# ── 툴 8: 단군에게 일반 질의 (피지수→단군) ───────────────────────────────────

@mcp.tool()
def physis_ask_dangun(question: str) -> str:
    """
    피지수가 본영 단군에게 일반 질의를 보낸다.
    에스컬레이션이 아닌 일반 판단·자문 요청에 사용.

    Args:
        question: 단군에게 전달할 질문 또는 요청

    Returns:
        단군 정반합 응답 텍스트
    """
    if not question.strip():
        return "[오류] question은 비어있을 수 없습니다."
    return _call_dangun_brain(question.strip())


# ── 툴 9: POPEYEs 현장 데이터 학습 ──────────────────────────────────────────

@mcp.tool()
def physis_learn_from_popeys(date: str) -> dict:
    """
    피지수가 POPEYEs 현장 데이터를 학습한다.
    지정 날짜의 TeamReport + MasterReport를 읽어 HyperRAG(ChromaVectorIndex)에 저장한다.

    Args:
        date: 학습할 날짜 (YYYY-MM-DD)

    Returns:
        {"learned": int, "date": str, "summary": str} 또는 {"error": str}
    """
    import asyncio
    from scripts.turso_reader import fetch_popeys_daily
    from scripts.embedding import ChromaVectorIndex, build_default_embedder
    from scripts.cma import memory_store

    if not date.strip():
        return {"error": "date는 비어있을 수 없습니다."}

    # 1. POPEYEs 데이터 읽기
    try:
        data = asyncio.run(fetch_popeys_daily(date.strip()))
    except Exception as e:
        return {"error": f"Turso 읽기 실패: {e}"}

    if not data.get("team_reports"):
        return {"error": f"{date} 데이터 없음"}

    # 2. HyperRAG에 학습 (CMA memory_store)
    idx = ChromaVectorIndex()
    emb = build_default_embedder()
    learned = 0

    for report in data["team_reports"]:
        content = f"[{date}][{report['team']}] 작업: {report['content']} / 인원: {report['worker_count']}명"
        result = memory_store(content, "popeys_daily", idx, emb)
        if result["status"] in ("stored", "merged"):
            learned += 1

    # 마스터 요약도 저장
    if data.get("master_summary"):
        memory_store(f"[{date}][마스터일보] {data['master_summary']}", "popeys_master", idx, emb)

    return {
        "learned": learned,
        "date": date,
        "summary": f"{date} 팀별 보고 {len(data['team_reports'])}건, {learned}건 학습 완료 (총 {data['total_workers']}명)"
    }


# ── 피지수 Vault 도구 ─────────────────────────────────────────────────────────

VAULT_ROOT = _FORGE_ROOT / "physis_memory"
_VAULT_WIKI = VAULT_ROOT / "wiki"
_VAULT_SCRIPTS = VAULT_ROOT / "scripts"


@mcp.tool()
def vault_query(question: str, n: int = 3) -> dict:
    """피지수 뇌에서 질문과 유사한 지식 검색 (Hot Tier 백링크 + Cold Tier 벡터)

    Args:
        question: 검색할 질문 또는 키워드
        n: 반환할 최대 결과 수

    Returns:
        {"hot": [...], "cold": [...]} — Hot(Vault 노트) + Cold(ChromaDB 데자뷔) 결과
    """
    import subprocess, json
    hot_results = []
    cold_results = []

    # Hot Tier: Graphify BFS 탐색
    try:
        r = subprocess.run(
            ["graphify", "query", question, "--budget", "1000",
             "--graph", str(VAULT_ROOT / "graphify-out" / "graph.json")],
            capture_output=True, text=True, cwd=str(VAULT_ROOT)
        )
        if r.returncode == 0 and r.stdout.strip():
            hot_results = [{"source": "hot_tier", "content": r.stdout[:1000]}]
    except Exception as e:
        hot_results = [{"source": "hot_tier", "error": str(e)}]

    # Cold Tier: ChromaDB 벡터 검색
    try:
        sys.path.insert(0, str(_VAULT_SCRIPTS))
        from cold_tier import recall
        hits = recall(question, n)
        cold_results = [{"source": "cold_tier", **h} for h in hits]
    except Exception as e:
        cold_results = [{"source": "cold_tier", "error": str(e)}]

    return {"hot": hot_results, "cold": cold_results, "query": question}


@mcp.tool()
def vault_ingest(title: str, content: str, tags: str = "") -> dict:
    """피지수 뇌에 새 지식 노드 추가 (wiki/ 폴더에 마크다운 생성)

    Args:
        title: 노트 제목 (파일명이 됨)
        content: 노트 본문 (마크다운, [[백링크]] 포함 가능)
        tags: 쉼표 구분 태그 (선택)

    Returns:
        {"path": str, "status": "created" | "updated"}
    """
    from datetime import datetime
    safe_title = title.replace("/", "_").replace("\\", "_")
    path = _VAULT_WIKI / f"{safe_title}.md"
    tag_list = [t.strip() for t in tags.split(",") if t.strip()]
    now = datetime.now().strftime("%Y-%m-%d")

    note = f"""---
type: wiki
created: {now}
tags: [{", ".join(tag_list)}]
ref_count: 0
outcome_score: 0.0
---

# {title}

{content}

## 연결

- [[홍익인간]]
"""
    status = "updated" if path.exists() else "created"
    path.write_text(note, encoding="utf-8")

    # log.md 업데이트
    log_path = _VAULT_WIKI / "log.md"
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"\n- [{ts}] vault_ingest: [[{safe_title}]] {status}")

    return {"path": str(path), "status": status, "title": title}


@mcp.tool()
def vault_cycle() -> dict:
    """피지수 사이클 실행 — CMA심사→소급평가→망각·승격→Graphify 갱신

    Returns:
        {"status": "ok", "graph": {"nodes": int, "edges": int}}
    """
    import subprocess
    r = subprocess.run(
        [sys.executable, str(_VAULT_SCRIPTS / "physis_cycle.py")],
        capture_output=True, text=True, cwd=str(_VAULT_SCRIPTS),
        env={**os.environ, "ANTHROPIC_API_KEY": os.environ.get("ANTHROPIC_API_KEY", "")}
    )

    # 그래프 통계 파싱
    nodes, edges = 0, 0
    for line in r.stdout.splitlines():
        if "nodes" in line and "edges" in line:
            import re
            m = re.search(r"(\d+) nodes.*?(\d+) edges", line)
            if m:
                nodes, edges = int(m.group(1)), int(m.group(2))

    return {
        "status": "ok" if r.returncode == 0 else "error",
        "graph": {"nodes": nodes, "edges": edges},
        "log": r.stdout[-500:] if r.stdout else ""
    }


@mcp.tool()
def vault_status() -> dict:
    """피지수 뇌 현황 — 노트 수, God Node 수, Cold Tier 규모

    Returns:
        {"total_notes": int, "god_nodes": int, "cold_tier": int, "graph": dict}
    """
    import json

    total = len(list(VAULT_ROOT.glob("**/*.md")))
    god_nodes = len(list((VAULT_ROOT / "god_nodes").glob("*.md")))
    wiki_notes = len(list(_VAULT_WIKI.glob("*.md")))

    # ChromaDB Cold Tier 크기
    cold_count = 0
    try:
        sys.path.insert(0, str(_VAULT_SCRIPTS))
        from cold_tier import stats
        cold_count = stats()
    except Exception:
        pass

    # Graphify 그래프 통계
    graph_json = VAULT_ROOT / "graphify-out" / "graph.json"
    graph = {}
    if graph_json.exists():
        try:
            data = json.loads(graph_json.read_text())
            graph = {"nodes": len(data.get("nodes", [])), "edges": len(data.get("edges", []))}
        except Exception:
            pass

    return {
        "total_notes": total,
        "god_nodes": god_nodes,
        "wiki_notes": wiki_notes,
        "cold_tier": cold_count,
        "graph": graph
    }


# ── 진입점 ────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    mcp.run(transport="stdio")
