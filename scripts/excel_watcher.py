"""
엑셀 감시 학습기 — excel_watcher.py
watch/ 폴더의 엑셀 파일을 파싱해 피지수 ChromaDB에 저장.
이미 처리한 파일은 processed/ 로 이동 (재처리 방지).
"""
from __future__ import annotations

import hashlib
import json
import sys
import time
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

WATCH_DIR = ROOT / "watch"
PROCESSED_DIR = ROOT / "watch" / "processed"
PROCESSED_LOG = ROOT / "watch" / ".processed_log.json"

EXCEL_EXTENSIONS = {".xlsx", ".xls"}

# ---------------------------------------------------------------------------
# 해시 / 로그 유틸
# ---------------------------------------------------------------------------

def _file_hash(filepath: Path) -> str:
    """파일 SHA-256 해시 반환 (재처리 방지용)."""
    sha = hashlib.sha256()
    with filepath.open("rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            sha.update(chunk)
    return sha.hexdigest()


def _load_processed_log() -> dict:
    """처리된 파일 해시 목록을 로드한다. 파일이 없으면 빈 dict 반환."""
    if PROCESSED_LOG.exists():
        try:
            return json.loads(PROCESSED_LOG.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def _save_processed_log(log: dict) -> None:
    """처리된 파일 해시 목록을 저장한다."""
    PROCESSED_LOG.write_text(json.dumps(log, ensure_ascii=False, indent=2), encoding="utf-8")


# ---------------------------------------------------------------------------
# 엑셀 파싱
# ---------------------------------------------------------------------------

def parse_excel(filepath: Path) -> list[str]:
    """엑셀 파일에서 의미 있는 텍스트 청크를 추출한다."""
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    chunks: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text: list[str] = []

        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                rows_text.append(" | ".join(cells))

        if rows_text:
            # 50행 단위로 청크
            for i in range(0, len(rows_text), 50):
                chunk = (
                    f"[파일:{filepath.name}][시트:{sheet_name}]\n"
                    + "\n".join(rows_text[i : i + 50])
                )
                chunks.append(chunk)

    wb.close()
    return chunks


# ---------------------------------------------------------------------------
# 단일 파일 처리
# ---------------------------------------------------------------------------

def process_file(filepath: Path, index, embedder) -> dict:
    """단일 파일을 파싱해 ChromaDB에 저장한 뒤 processed/ 로 이동한다."""
    from scripts.titans_memory import store_memory

    chunks = parse_excel(filepath)
    stored = 0

    for chunk in chunks:
        if store_memory(chunk, f"excel:{filepath.stem}", index, embedder):
            stored += 1

    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    dest = PROCESSED_DIR / filepath.name
    # 동일 이름이 이미 있으면 타임스탬프 접미사 추가
    if dest.exists():
        ts = int(time.time())
        dest = PROCESSED_DIR / f"{filepath.stem}_{ts}{filepath.suffix}"
    filepath.rename(dest)

    return {"file": filepath.name, "chunks": len(chunks), "stored": stored}


# ---------------------------------------------------------------------------
# 1회 스캔
# ---------------------------------------------------------------------------

def run_once() -> None:
    """watch/ 폴더를 1회 스캔해 새 엑셀 파일을 처리한다."""
    # openpyxl 설치 확인
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("openpyxl이 설치되어 있지 않습니다.")
        print("설치 명령: pip install openpyxl")
        sys.exit(1)

    # watch/ 폴더 자동 생성
    WATCH_DIR.mkdir(parents=True, exist_ok=True)

    excel_files = [
        f for f in WATCH_DIR.iterdir()
        if f.is_file() and f.suffix.lower() in EXCEL_EXTENSIONS
    ]

    if not excel_files:
        print("watch/ 폴더 스캔 완료 (파일 없음)")
        return

    # ChromaDB + 임베더 초기화
    from dotenv import load_dotenv
    load_dotenv()

    from scripts.embedding import ChromaVectorIndex, build_default_embedder
    index = ChromaVectorIndex()
    embedder = build_default_embedder()

    processed_log = _load_processed_log()
    processed_count = 0

    for filepath in excel_files:
        file_hash = _file_hash(filepath)

        if file_hash in processed_log:
            print(f"[SKIP] {filepath.name} — 이미 처리됨")
            continue

        print(f"[처리] {filepath.name} ...")
        try:
            result = process_file(filepath, index, embedder)
            processed_log[file_hash] = {
                "file": result["file"],
                "chunks": result["chunks"],
                "stored": result["stored"],
            }
            _save_processed_log(processed_log)
            print(
                f"[완료] {result['file']} — "
                f"청크 {result['chunks']}개 중 {result['stored']}개 저장됨"
            )
            processed_count += 1
        except Exception as exc:
            print(f"[오류] {filepath.name} 처리 실패: {exc}")

    print(f"watch/ 폴더 스캔 완료 ({processed_count}개 처리됨)")


# ---------------------------------------------------------------------------
# 반복 감시
# ---------------------------------------------------------------------------

def run_watch(interval: int = 60) -> None:
    """interval초마다 반복 스캔한다."""
    print(f"엑셀 감시 시작 (간격: {interval}초). 종료하려면 Ctrl+C.")
    while True:
        try:
            run_once()
        except KeyboardInterrupt:
            print("\n감시 종료.")
            sys.exit(0)
        except Exception as exc:
            print(f"[경고] 스캔 중 예외 발생: {exc}")
        time.sleep(interval)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if "--once" in sys.argv:
        run_once()
    else:
        run_watch(interval=60)
