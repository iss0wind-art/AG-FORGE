"""
동우건설 공사일보 Excel 생성기.
4블록 병렬 레이아웃 (50명 단위) — prompt_plan.md 역설계 스펙 기반.
저장 경로: Documents/신고조선/H2OWIND지국/공사일보/YYYY-MM-DD_공사일보.xlsx
"""
import os
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# 저장 루트 — 환경변수로 오버라이드 가능
_DEFAULT_OUTPUT = Path.home() / "Documents" / "신고조선" / "H2OWIND지국" / "공사일보"
OUTPUT_DIR = Path(os.getenv("REPORT_OUTPUT_DIR", str(_DEFAULT_OUTPUT)))

SITE_NAME = "에코델타시티 24BL 건립공사 현장"
BLOCK_SIZE = 50  # 블록당 최대 인원


def _thin_border() -> Border:
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_fill() -> PatternFill:
    return PatternFill("solid", fgColor="D9E1F2")


def _write_block_header(ws, col_start: int, row: int) -> None:
    """블록 헤더(순번|직종|성명|공수|작업내용) 1행 기입."""
    headers = ["순번", "직종", "성명", "공수", "작업내용"]
    widths = [6, 8, 10, 6, 20]
    for i, (h, w) in enumerate(zip(headers, widths)):
        col = col_start + i
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
        cell.fill = _header_fill()
        ws.column_dimensions[get_column_letter(col)].width = w


def _write_worker_row(ws, col_start: int, row: int, seq: int, worker: dict) -> None:
    """개인 데이터 1행 기입."""
    values = [
        seq,
        worker.get("job_type", ""),
        worker.get("worker_name", ""),
        worker.get("manday", 1.0),
        worker.get("work_content", ""),
    ]
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=col_start + i, value=v)
        cell.border = _thin_border()
        cell.font = Font(size=9)
        cell.alignment = Alignment(horizontal="center" if i != 4 else "left", vertical="center")


def generate_공사일보(date: str, workers: list[dict], team_reports: list[dict]) -> Path:
    """
    공사일보 Excel을 생성하고 저장 경로를 반환한다.

    date: 'YYYY-MM-DD'
    workers: [{ team, job_type, worker_name, manday, work_content, section }]
    team_reports: [{ team, manual_worker_count }]
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{date} 공사일보"
    ws.sheet_view.showGridLines = False

    # --- 헤더 섹션 (행 1~9) ---
    dt = datetime.strptime(date, "%Y-%m-%d")

    ws.merge_cells("A1:T1")
    title_cell = ws["A1"]
    title_cell.value = f"일  일  공  사  일  보    ({date})"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:T2")
    ws["A2"].value = f"현장명: {SITE_NAME}"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    # 팀별 인원 집계 헤더
    team_count_map = {r["team"]: r.get("manual_worker_count", 0) for r in team_reports}
    total = sum(team_count_map.values()) or len(workers)

    ws.merge_cells("A3:T3")
    summary_parts = [f"{t}: {c}명" for t, c in team_count_map.items() if c > 0]
    ws["A3"].value = f"총 {total}명 — " + ", ".join(summary_parts[:10])
    ws["A3"].font = Font(size=9)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 16

    # 빈 여백 행
    for r in range(4, 9):
        ws.row_dimensions[r].height = 14

    # --- 블록 헤더 (행 9) ---
    HEADER_ROW = 9
    block_col_starts = [1, 6, 11, 16]  # 4블록: A, F, K, P 시작
    for col_start in block_col_starts:
        _write_block_header(ws, col_start, HEADER_ROW)

    # --- 개인 명세 섹션 (행 10~) ---
    DATA_START_ROW = 10
    blocks: list[list[dict]] = [
        workers[i * BLOCK_SIZE: (i + 1) * BLOCK_SIZE]
        for i in range(4)
    ]

    max_rows = max((len(b) for b in blocks), default=0)

    for row_offset in range(max_rows):
        row = DATA_START_ROW + row_offset
        ws.row_dimensions[row].height = 15
        for block_idx, col_start in enumerate(block_col_starts):
            block = blocks[block_idx]
            if row_offset < len(block):
                seq = block_idx * BLOCK_SIZE + row_offset + 1
                _write_worker_row(ws, col_start, row, seq, block[row_offset])
            else:
                # 빈 셀도 테두리 유지
                for i in range(5):
                    ws.cell(row=row, column=col_start + i).border = _thin_border()

    # --- 저장 ---
    filename = f"{date}_공사일보.xlsx"
    filepath = OUTPUT_DIR / filename
    wb.save(filepath)
    return filepath
